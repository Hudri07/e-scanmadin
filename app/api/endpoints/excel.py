import pandas as pd
import io
from typing import List
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Definisikan Router
router = APIRouter()

@router.post("/ekspor-excel")
async def ekspor_excel(data: List[dict]):
    if not data:
        raise HTTPException(status_code=400, detail="Tidak ada data untuk diekspor")

    try:
        df_raw = pd.DataFrame(data)

        # Pivot data agar mapel jadi kolom
        df_pivot = df_raw.pivot(index=['nomor_peserta', 'nama'], 
                                columns='mapel', 
                                values='skor').reset_index()

        list_mapel = [col for col in df_pivot.columns if col not in ['nomor_peserta', 'nama']]
        jumlah_mapel = len(list_mapel)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Tulis data mulai baris ke-2
            df_pivot.to_excel(writer, index=False, sheet_name='Rekap Nilai', startrow=1)
            worksheet = writer.sheets['Rekap Nilai']
            
            # Style untuk Header
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center')

            # Menggabungkan baris 1 dan 2 untuk kolom A dan B
            worksheet.merge_cells('A1:A2')
            worksheet['A1'] = 'Nomor Peserta'
            
            worksheet.merge_cells('B1:B2')
            worksheet['B1'] = 'Nama Siswa'

            # AUTO-FIT UNTUK KOLOM MAPEL (Mulai dari kolom C dan seterusnya)
            for i, mapel in enumerate(list_mapel, start=3): # start=3 artinya kolom C
                kolom_letter = get_column_letter(i)
                
                # Hitung panjang teks (judul mapel)
                # beri sedikit padding agar tidak terlalu mepet
                panjang_teks = len(str(mapel)) +  10
                
                # Atur lebar kolom berdasarkan panjang judul mapel
                worksheet.column_dimensions[kolom_letter].width = panjang_teks

            # MERGE HORIZONTAL (Mata Pelajaran)
            if jumlah_mapel > 0:
                kolom_awal_mapel = "C"
                kolom_akhir_mapel = get_column_letter(2 + jumlah_mapel)
                
                range_merge_mapel = f"{kolom_awal_mapel}1:{kolom_akhir_mapel}1"
                worksheet.merge_cells(range_merge_mapel)
                worksheet[f"{kolom_awal_mapel}1"] = 'Mata Pelajaran'

            # Atur Lebar Kolom & Bold Header
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 35
            
            # Terapkan Bold dan Center ke semua sel di baris 1 dan 2
            for row in worksheet.iter_rows(min_row=1, max_row=2):
                for cell in row:
                    cell.font = bold_font
                    cell.alignment = center_align

        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=rekap_nilai_madin.xlsx"}
        )
    except Exception as e:
        print(f"Error: {str(e)}")
        raise HTTPException(status_code=500, detail="Gagal generate excel")